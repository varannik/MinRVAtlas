"""
PDF / Excel quality report export — Task-33.
Routes:
  GET  /api/v1/reports/export/{project_id}?format=pdf|xlsx
  POST /api/v1/reports/schedule-digest           wire into digest scheduler
"""
import io
import logging
from datetime import datetime
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models import DQARule, DQARun, DQAViolation, Project, ProjectMember

logger = logging.getLogger("datasentinel.reports")
router = APIRouter()


# ── Access guard ──────────────────────────────────────────────────────────────

def _require_access(db: Session, project_id, user):
    role = getattr(user, "role", "")
    if role in ("admin", "super_admin"):
        return
    member = db.query(ProjectMember).filter(
        ProjectMember.project_id == project_id,
        ProjectMember.user_id == user.id,
    ).first()
    if not member:
        raise HTTPException(403, "You don't have access to this project")


# ── Data assembly ─────────────────────────────────────────────────────────────

def _assemble_report_data(project_id: str, db: Session) -> dict:
    """Build a dict of all data needed for both PDF and Excel reports."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    runs = (
        db.query(DQARun)
        .filter(DQARun.project_id == project_id, DQARun.status == "completed")
        .order_by(DQARun.triggered_at.desc())
        .limit(20)
        .all()
    )
    recent_run = runs[0] if runs else None
    violations = []
    if recent_run:
        violations = (
            db.query(DQAViolation)
            .filter(DQAViolation.run_id == recent_run.id)
            .limit(1000)
            .all()
        )

    rules = db.query(DQARule).filter(DQARule.project_id == project_id).all()

    # Aggregations
    by_severity = {s: sum(1 for v in violations if v.severity == s)
                   for s in ["critical", "high", "medium", "low"]}
    by_dimension = {}
    for v in violations:
        by_dimension[v.dimension] = by_dimension.get(v.dimension, 0) + 1

    run_history = [
        {
            "date":           r.triggered_at.strftime("%Y-%m-%d %H:%M") if r.triggered_at else "",
            "readiness":      round((r.readiness_score or 0) * 100, 1),
            "violations":     r.total_violations or 0,
            "gate_passed":    "PASSED" if r.gate_passed else "FAILED",
            "rules_executed": r.rules_executed or 0,
        }
        for r in runs
    ]

    readiness_trend = [round((r.readiness_score or 0) * 100, 1) for r in reversed(runs)]
    avg_readiness = round(sum(readiness_trend) / len(readiness_trend), 1) if readiness_trend else 0

    return {
        "project":       project,
        "project_name":  project.name,
        "generated_at":  datetime.utcnow(),
        "recent_run":    recent_run,
        "violations":    violations,
        "rules":         rules,
        "run_history":   run_history,
        "by_severity":   by_severity,
        "by_dimension":  by_dimension,
        "avg_readiness": avg_readiness,
        "readiness_trend": readiness_trend,
        "total_runs":    len(runs),
    }


# ── PDF builder ───────────────────────────────────────────────────────────────

def _build_pdf(data: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=0.75 * inch, leftMargin=0.75 * inch,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    GREEN  = colors.HexColor("#27ae60")
    RED    = colors.HexColor("#e74c3c")
    AMBER  = colors.HexColor("#f39c12")
    DARK   = colors.HexColor("#1a1a2e")
    LIGHT  = colors.HexColor("#f0f4f8")

    h1 = ParagraphStyle("H1", parent=styles["Heading1"], textColor=DARK, fontSize=20, spaceAfter=6)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=DARK, fontSize=14, spaceBefore=14, spaceAfter=4)
    body = ParagraphStyle("Body", parent=styles["Normal"], fontSize=9, leading=13)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

    story = []

    # ── Title ──────────────────────────────────────────────────────────────────
    story.append(Paragraph("DataSentinel DQA Report", h1))
    story.append(Paragraph(f"Project: <b>{data['project_name']}</b>", body))
    story.append(Paragraph(f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M')} UTC", small))
    story.append(HRFlowable(width="100%", thickness=1, color=DARK))
    story.append(Spacer(1, 0.15 * inch))

    # ── Summary box ────────────────────────────────────────────────────────────
    rr = data["recent_run"]
    if rr:
        score_pct = round((rr.readiness_score or 0) * 100, 1)
        gate_color = GREEN if rr.gate_passed else RED
        gate_label = "PASSED" if rr.gate_passed else "FAILED"
        summary_data = [
            ["Metric", "Value"],
            ["Latest Readiness Score", f"{score_pct}%"],
            ["Gate Status",            gate_label],
            ["Total Violations",       str(rr.total_violations or 0)],
            ["Rules Executed",         str(rr.rules_executed or 0)],
            ["Avg Readiness (last 20)", f"{data['avg_readiness']}%"],
            ["Total Completed Runs",   str(data["total_runs"])],
        ]
        tbl = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",     (0, 0), (-1, 0), 9),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND",   (0, 1), (-1, -1), LIGHT),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("FONTSIZE",     (0, 1), (-1, -1), 9),
            ("LEFTPADDING",  (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        story.append(Paragraph("Executive Summary", h2))
        story.append(tbl)
        story.append(Spacer(1, 0.15 * inch))

    # ── Violations by severity ─────────────────────────────────────────────────
    story.append(Paragraph("Violations by Severity", h2))
    sev_data = [["Severity", "Count"]] + [
        [s.capitalize(), str(data["by_severity"].get(s, 0))]
        for s in ["critical", "high", "medium", "low"]
    ]
    sev_tbl = Table(sev_data, colWidths=[2.5 * inch, 1.5 * inch])
    sev_colors_map = {"Critical": RED, "High": AMBER, "Medium": colors.HexColor("#3498db"), "Low": GREEN}
    sev_style = [
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("LEFTPADDING",(0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
    ]
    for i, sev in enumerate(["Critical", "High", "Medium", "Low"], start=1):
        sev_style.append(("TEXTCOLOR", (0, i), (0, i), sev_colors_map[sev]))
    sev_tbl.setStyle(TableStyle(sev_style))
    story.append(sev_tbl)
    story.append(Spacer(1, 0.15 * inch))

    # ── Top violations table ───────────────────────────────────────────────────
    if data["violations"]:
        story.append(Paragraph("Top Violations (Latest Run)", h2))
        viol_data = [["Rule", "Dimension", "Severity", "Field", "Count"]]
        for v in sorted(data["violations"], key=lambda x: -getattr(x, "record_count", 0))[:15]:
            viol_data.append([
                (v.rule_name or "")[:30],
                v.dimension or "",
                v.severity or "",
                (v.affected_field or "")[:25],
                str(v.record_count or 0),
            ])
        col_widths = [2.2 * inch, 1.2 * inch, 0.9 * inch, 1.5 * inch, 0.7 * inch]
        vtbl = Table(viol_data, colWidths=col_widths)
        vtbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("LEFTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("WORDWRAP",     (0, 0), (-1, -1), "CJK"),
        ]))
        story.append(vtbl)
        story.append(Spacer(1, 0.15 * inch))

    # ── Run history ────────────────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Run History (Last 20 Runs)", h2))
    hist_data = [["Date", "Readiness %", "Violations", "Gate", "Rules"]]
    for r in data["run_history"]:
        hist_data.append([
            r["date"], f"{r['readiness']}%",
            str(r["violations"]), r["gate_passed"], str(r["rules_executed"]),
        ])
    htbl = Table(hist_data, colWidths=[1.8 * inch, 1.2 * inch, 1.1 * inch, 1.0 * inch, 0.9 * inch])
    htbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]))
    story.append(htbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(data: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _write_header(ws, headers: list[str]):
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

    def _autosize(ws):
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: Summary ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    rr = data["recent_run"]
    ws1.append(["DataSentinel DQA Report"])
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append([f"Project: {data['project_name']}"])
    ws1.append([f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M')} UTC"])
    ws1.append([])
    _write_header(ws1, ["Metric", "Value"])
    if rr:
        for row in [
            ("Latest Readiness Score", f"{round((rr.readiness_score or 0)*100,1)}%"),
            ("Gate Status",            "PASSED" if rr.gate_passed else "FAILED"),
            ("Total Violations",       rr.total_violations or 0),
            ("Rules Executed",         rr.rules_executed or 0),
            ("Average Readiness",      f"{data['avg_readiness']}%"),
            ("Total Completed Runs",   data["total_runs"]),
        ]:
            ws1.append(list(row))
    _autosize(ws1)

    # ── Sheet 2: Violations ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Violations")
    _write_header(ws2, ["Rule ID", "Rule Name", "Dimension", "Severity", "Affected Field", "Record Count", "Status"])
    for i, v in enumerate(data["violations"], start=2):
        ws2.append([
            v.rule_id, v.rule_name, v.dimension, v.severity,
            v.affected_field, v.record_count or 0, v.status,
        ])
        if i % 2 == 0:
            for col in range(1, 8):
                ws2.cell(i, col).fill = alt_fill
    _autosize(ws2)

    # ── Sheet 3: Rules ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Rules")
    _write_header(ws3, ["Rule ID", "Name", "Dimension", "Severity", "Hard Gate", "Weight", "Active"])
    for i, r in enumerate(data["rules"], start=2):
        ws3.append([
            r.rule_id, r.rule_name, r.dimension, r.severity,
            "Yes" if r.is_hard_gate else "No",
            round(r.weight or 0, 3),
            "Yes" if r.is_active else "No",
        ])
        if i % 2 == 0:
            for col in range(1, 8):
                ws3.cell(i, col).fill = alt_fill
    _autosize(ws3)

    # ── Sheet 4: Run History ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Run History")
    _write_header(ws4, ["Date", "Readiness %", "Violations", "Gate", "Rules Executed"])
    for i, r in enumerate(data["run_history"], start=2):
        ws4.append([r["date"], r["readiness"], r["violations"], r["gate_passed"], r["rules_executed"]])
        if i % 2 == 0:
            for col in range(1, 6):
                ws4.cell(i, col).fill = alt_fill
    _autosize(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Export endpoint ───────────────────────────────────────────────────────────

@router.get("/export/{project_id}")
def export_report(
    project_id: UUID,
    format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Generate and stream a PDF or Excel quality report for a project."""
    _require_access(db, project_id, user)
    data = _assemble_report_data(str(project_id), db)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    safe_name = "".join(c if c.isalnum() else "_" for c in data["project_name"])[:40]
    filename = f"DataSentinel_{safe_name}_{ts}.{format}"

    if format == "pdf":
        try:
            content = _build_pdf(data)
        except Exception as exc:
            logger.error("PDF generation failed: %s", exc)
            raise HTTPException(500, f"PDF generation failed: {exc}")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    else:  # xlsx
        try:
            content = _build_excel(data)
        except Exception as exc:
            logger.error("Excel generation failed: %s", exc)
            raise HTTPException(500, f"Excel generation failed: {exc}")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


# ── Digest email with report attachment ──────────────────────────────────────

@router.post("/send-digest/{project_id}")
def send_report_digest(
    project_id: UUID,
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Send a PDF quality report by email. Body: {recipient_email, format?}."""
    _require_access(db, project_id, user)
    recipient = data.get("recipient_email", "")
    if not recipient:
        raise HTTPException(400, "recipient_email is required")

    fmt = data.get("format", "pdf")
    report_data = _assemble_report_data(str(project_id), db)
    ts = datetime.utcnow().strftime("%Y%m%d")
    safe_name = "".join(c if c.isalnum() else "_" for c in report_data["project_name"])[:30]

    try:
        if fmt == "pdf":
            attachment_bytes = _build_pdf(report_data)
            filename = f"DataSentinel_{safe_name}_{ts}.pdf"
            mime_type = "application/pdf"
        else:
            attachment_bytes = _build_excel(report_data)
            filename = f"DataSentinel_{safe_name}_{ts}.xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    except Exception as exc:
        raise HTTPException(500, f"Report generation failed: {exc}")

    # Send via SES
    try:
        import boto3

        from app.core.config import settings
        ses = boto3.client("ses", region_name=settings.AWS_REGION or "us-east-1")
        from email.mime.application import MIMEApplication
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        msg = MIMEMultipart("mixed")
        msg["Subject"] = f"DataSentinel Quality Report — {report_data['project_name']}"
        msg["From"]    = settings.ALERT_EMAIL_FROM or "noreply@datasentinel.io"
        msg["To"]      = recipient

        body = MIMEText(
            f"Please find attached the latest DQA quality report for "
            f"<b>{report_data['project_name']}</b>.<br><br>"
            f"Readiness Score: <b>{report_data['avg_readiness']}%</b><br>"
            f"Total Runs: {report_data['total_runs']}",
            "html",
        )
        msg.attach(body)
        att = MIMEApplication(attachment_bytes, _subtype="octet-stream")
        att.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(att)

        ses.send_raw_email(
            Source=msg["From"],
            Destinations=[recipient],
            RawMessage={"Data": msg.as_string()},
        )
        return {"ok": True, "message": f"Report sent to {recipient}"}
    except Exception as exc:
        logger.error("Report email failed: %s", exc)
        raise HTTPException(502, f"Email delivery failed: {str(exc)[:200]}")
