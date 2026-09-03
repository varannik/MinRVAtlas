from datetime import datetime, timedelta, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models import (
    DQARun,
    DQAViolation,
    ProjectMember,
    User,
    ViolationComment,
)
from app.schemas import (
    ViolationOut,
)

router = APIRouter()


def _require_project_access(db: Session, project_id, user) -> None:
    """Raise 403 if user is not admin/super_admin and not a member of the project."""
    role = getattr(user, "role", "")
    if role in ("admin", "super_admin"):
        return
    member = db.query(ProjectMember).filter(
        ProjectMember.project_id == project_id,
        ProjectMember.user_id == user.id,
    ).first()
    if not member:
        raise HTTPException(403, "You don't have access to this project")


def _get_violation_project_id(db: Session, v: DQAViolation):
    """Return the project_id for a violation (via its parent run)."""
    run = db.query(DQARun).filter(DQARun.id == v.run_id).first()
    return run.project_id if run else None


@router.get("/")
def list_violations(run_id: Optional[UUID] = None, dataset_id: Optional[UUID] = None,
                    severity: Optional[str] = None, status: Optional[str] = None,
                    dimension: Optional[str] = None,
                    offset: int = 0, limit: int = 100,
                    db: Session = Depends(get_db), user=Depends(get_current_user)):
    role = getattr(user, "role", "")
    q = db.query(DQAViolation)
    if run_id:
        # Verify the caller can access the run's project
        run = db.query(DQARun).filter(DQARun.id == run_id).first()
        if run:
            _require_project_access(db, run.project_id, user)
        q = q.filter(DQAViolation.run_id == run_id)
    elif role not in ("admin", "super_admin"):
        # Non-admin without run_id: restrict to projects the user belongs to
        project_ids = [pm.project_id for pm in db.query(ProjectMember).filter(
            ProjectMember.user_id == user.id
        ).all()]
        accessible_run_ids = [r.id for r in db.query(DQARun.id).filter(
            DQARun.project_id.in_(project_ids)
        ).all()]
        q = q.filter(DQAViolation.run_id.in_(accessible_run_ids))
    if dataset_id: q = q.filter(DQAViolation.dataset_id == dataset_id)
    if severity: q = q.filter(DQAViolation.severity == severity)
    if status: q = q.filter(DQAViolation.status == status)
    if dimension: q = q.filter(DQAViolation.dimension == dimension)
    limit = max(1, min(limit, 500))
    total = q.count()
    items = q.order_by(DQAViolation.created_at.desc()).offset(offset).limit(limit).all()
    return {"items": items, "total": total, "offset": offset, "limit": limit}

@router.get("/sla-overdue")
def sla_overdue_violations(
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Return violations that are past their SLA due date."""
    now = datetime.now(timezone.utc)
    q = db.query(DQAViolation).filter(
        DQAViolation.due_date < now,
        DQAViolation.status == "open",
    )
    if project_id:
        _require_project_access(db, project_id, user)
        run_ids = db.query(DQARun.id).filter(DQARun.project_id == project_id).all()
        run_ids = [r[0] for r in run_ids]
        q = q.filter(DQAViolation.run_id.in_(run_ids))
    else:
        # Non-admins: restrict to their accessible projects
        role = getattr(user, "role", "")
        if role not in ("admin", "super_admin"):
            project_ids = [pm.project_id for pm in db.query(ProjectMember).filter(
                ProjectMember.user_id == user.id
            ).all()]
            run_ids = [r.id for r in db.query(DQARun.id).filter(
                DQARun.project_id.in_(project_ids)
            ).all()]
            q = q.filter(DQAViolation.run_id.in_(run_ids))
    items = q.order_by(DQAViolation.due_date.asc()).limit(100).all()
    result = []
    for v in items:
        due = v.due_date
        if due and due.tzinfo is None:
            due = due.replace(tzinfo=timezone.utc)
        hours_overdue = int((now - due).total_seconds() / 3600) if due else 0
        result.append({
            "id": str(v.id), "rule_name": v.rule_name, "severity": v.severity,
            "dimension": v.dimension, "status": v.status,
            "due_date": v.due_date.isoformat() if v.due_date else None,
            "hours_overdue": hours_overdue,
            "assigned_to": str(v.assigned_to) if v.assigned_to else None,
        })
    return {"overdue": result, "count": len(result)}


@router.get("/{violation_id}", response_model=ViolationOut)
def get_violation(violation_id: UUID, db: Session = Depends(get_db), user=Depends(get_current_user)):
    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v: raise HTTPException(404, "Violation not found")
    project_id = _get_violation_project_id(db, v)
    if project_id:
        _require_project_access(db, project_id, user)
    return v

@router.get("/{violation_id}/ai-playbook")
async def ai_playbook(violation_id: UUID, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Generate a step-by-step AI remediation playbook for a violation."""
    from app.engines.ai.playbook_agent import generate_playbook
    from app.models import DQARun, KnowledgeBase, Project

    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v:
        raise HTTPException(404, "Violation not found")

    # Fetch KB entries for context
    run = db.query(DQARun).filter(DQARun.id == v.run_id).first()
    project_name = ""
    kb_entries = []
    if run:
        project = db.query(Project).filter(Project.id == run.project_id).first()
        if project:
            project_name = project.name
        try:
            from app.models import KnowledgeBase
            kb_entries = [
                {"parameter": k.parameter_name, "title": k.title, "content": k.content[:500]}
                for k in db.query(KnowledgeBase).filter(
                    KnowledgeBase.project_id == run.project_id
                ).limit(10).all()
            ]
        except Exception:
            pass

    result = await generate_playbook(v, kb_entries, project_name)
    return result


@router.patch("/{violation_id}/status")
def update_violation_status(violation_id: UUID, body: dict,
                             db: Session = Depends(get_db), user=Depends(get_current_user)):
    # F014: accept status in the request body, not as a query parameter
    status = (body.get("status") or "").strip()
    VALID_STATUSES = {"open", "acknowledged", "resolved", "false_positive", "waived"}
    if status not in VALID_STATUSES:
        raise HTTPException(400, f"Invalid status. Must be one of: {sorted(VALID_STATUSES)}")
    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v: raise HTTPException(404, "Violation not found")
    v.status = status; db.commit()
    return {"message": "Status updated", "status": status}


@router.patch("/bulk-status")
def bulk_update_status(
    body: dict,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """B1-#3: Update status on multiple violations in one call.
    Body: { ids: [uuid, ...], status: "acknowledged" | "resolved" | ... }
    """
    ids = body.get("ids") or []
    status = (body.get("status") or "").strip()
    VALID_STATUSES = {"open", "acknowledged", "resolved", "false_positive", "waived"}
    if not ids:
        raise HTTPException(400, "ids list is required")
    if status not in VALID_STATUSES:
        raise HTTPException(400, f"Invalid status. Must be one of: {sorted(VALID_STATUSES)}")
    updated = (
        db.query(DQAViolation)
        .filter(DQAViolation.id.in_(ids))
        .all()
    )
    for v in updated:
        v.status = status
    db.commit()
    return {"updated": len(updated), "status": status}


@router.get("/export.xlsx")
def export_violations_xlsx(
    severity: Optional[str] = None,
    status: Optional[str] = None,
    dimension: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """B1-#7: Export violations as a proper Excel workbook."""
    import io

    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Font, PatternFill

    q = db.query(DQAViolation)
    if severity: q = q.filter(DQAViolation.severity == severity)
    if status:   q = q.filter(DQAViolation.status == status)
    if dimension: q = q.filter(DQAViolation.dimension == dimension)
    violations = q.order_by(DQAViolation.created_at.desc()).limit(5000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations"

    headers = ["ID", "Rule ID", "Rule Name", "Dimension", "Severity",
               "Affected Field", "Record Count", "Status", "Created At"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, v in enumerate(violations, 2):
        ws.cell(row=row_idx, column=1, value=str(v.id))
        ws.cell(row=row_idx, column=2, value=v.rule_id)
        ws.cell(row=row_idx, column=3, value=v.rule_name)
        ws.cell(row=row_idx, column=4, value=v.dimension)
        ws.cell(row=row_idx, column=5, value=v.severity)
        ws.cell(row=row_idx, column=6, value=v.affected_field or "")
        ws.cell(row=row_idx, column=7, value=v.record_count)
        ws.cell(row=row_idx, column=8, value=v.status)
        ws.cell(row=row_idx, column=9, value=v.created_at.isoformat() if v.created_at else "")

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=violations.xlsx"},
    )


# ── Phase 3: Assignment & SLA ─────────────────────────────────────────────────

@router.patch("/{violation_id}/assign")
def assign_violation(
    violation_id: UUID,
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Assign violation to a user + set SLA due date."""
    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v:
        raise HTTPException(404, "Violation not found")

    if "assigned_to" in data:
        v.assigned_to = data["assigned_to"] or None
    if "sla_hours" in data and data["sla_hours"]:
        v.sla_hours = int(data["sla_hours"])
        v.due_date = datetime.now(timezone.utc) + timedelta(hours=int(data["sla_hours"]))
    if "due_date" in data and data["due_date"]:
        v.due_date = data["due_date"]

    db.commit()

    # Look up assigned user name
    assignee_name = None
    if v.assigned_to:
        u = db.query(User).filter(User.id == v.assigned_to).first()
        assignee_name = u.full_name if u else None

    return {
        "id":            str(v.id),
        "assigned_to":   str(v.assigned_to) if v.assigned_to else None,
        "assignee_name": assignee_name,
        "due_date":      v.due_date.isoformat() if v.due_date else None,
        "sla_hours":     v.sla_hours,
    }


# ── Phase 3: Comments ─────────────────────────────────────────────────────────

@router.get("/{violation_id}/comments")
def list_comments(
    violation_id: UUID,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v:
        raise HTTPException(404, "Violation not found")
    from sqlalchemy.orm import joinedload
    comments = (
        db.query(ViolationComment)
        .options(joinedload(ViolationComment.user))
        .filter(ViolationComment.violation_id == violation_id)
        .order_by(ViolationComment.created_at.asc())
        .all()
    )
    result = []
    for c in comments:
        result.append({
            "id":         str(c.id),
            "user_id":    str(c.user_id),
            "user_name":  c.user.full_name if c.user else "Unknown",
            "message":    c.message,
            "created_at": c.created_at.isoformat() if c.created_at else None,
        })
    return {"comments": result, "count": len(result)}


@router.post("/{violation_id}/comments")
def add_comment(
    violation_id: UUID,
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v:
        raise HTTPException(404, "Violation not found")
    msg = (data.get("message") or "").strip()
    if not msg:
        raise HTTPException(400, "Message cannot be empty")
    # F018: cap comment length to prevent database bloat
    if len(msg) > 10_000:
        raise HTTPException(400, "Comment exceeds the maximum length of 10,000 characters")
    c = ViolationComment(violation_id=violation_id, user_id=user.id, message=msg)
    db.add(c)
    db.commit()
    db.refresh(c)
    return {
        "id":         str(c.id),
        "user_id":    str(c.user_id),
        "user_name":  user.full_name,
        "message":    c.message,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


@router.get("/{violation_id}/credit-impact")
def violation_credit_impact(
    violation_id: UUID,
    spot_price_per_tonne: float = 50.0,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Estimate the carbon credit financial impact of a violation.
    spot_price_per_tonne: £/t CO₂ (default 50 — Puro.Earth indicative price).
    """
    v = db.query(DQAViolation).filter(DQAViolation.id == violation_id).first()
    if not v:
        raise HTTPException(404, "Violation not found")

    affected_rows = v.record_count or len(v.affected_rows or [])

    # Physical defaults for CCS sensor data
    freq_minutes     = 2      # 2-min sampling (STR1 standard)
    avg_rate_m3h     = 87.0   # typical CO₂ injection rate (seeded from I-04 test data)
    co2_density_t_m3 = 0.71   # t/m³ supercritical CO₂ at injection conditions

    # Refine rate estimate from violation detail if available
    detail = v.violation_detail or {}
    if isinstance(detail, dict):
        if "mean" in detail:
            try: avg_rate_m3h = max(0.0, float(detail["mean"]))
            except (TypeError, ValueError): pass
        elif "values" in detail and detail["values"]:
            try:
                vals = [float(x) for x in detail["values"] if x is not None]
                if vals: avg_rate_m3h = sum(vals) / len(vals)
            except (TypeError, ValueError): pass

    hours_affected  = (affected_rows * freq_minutes) / 60.0
    co2_m3          = avg_rate_m3h * hours_affected
    co2_tonnes      = round(co2_m3 * co2_density_t_m3, 3)

    # Severity discount: critical violations may fully invalidate readings;
    # lower severities represent partial quality degradation risk.
    severity_factor = {"critical": 1.0, "high": 0.75, "medium": 0.50, "low": 0.25}.get(
        v.severity or "medium", 0.50
    )
    at_risk_tonnes = round(co2_tonnes * severity_factor, 3)
    market_value   = round(at_risk_tonnes * spot_price_per_tonne, 2)

    return {
        "violation_id":        str(violation_id),
        "rule_id":             v.rule_id,
        "dimension":           v.dimension,
        "severity":            v.severity,
        "affected_rows":       affected_rows,
        "estimated_co2_tonnes": co2_tonnes,
        "at_risk_tonnes":      at_risk_tonnes,
        "severity_factor":     severity_factor,
        "market_value":        market_value,
        "currency":            "GBP",
        "spot_price_per_tonne": spot_price_per_tonne,
        "assumptions": {
            "frequency_minutes":       freq_minutes,
            "avg_injection_rate_m3h":  round(avg_rate_m3h, 2),
            "co2_density_t_m3":        co2_density_t_m3,
            "hours_affected":          round(hours_affected, 3),
        },
    }
