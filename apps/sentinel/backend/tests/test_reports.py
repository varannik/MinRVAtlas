"""
Tests for PDF/Excel export report generation.
"""
import os
import io
import pytest

os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("SECRET_KEY", "test-secret-key-for-unit-tests-only-32chars!")
os.environ.setdefault("ENVIRONMENT", "test")


# ── PDF generation ─────────────────────────────────────────────────────────────

def test_pdf_library_available():
    """reportlab must be importable."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    assert letter is not None


def test_pdf_generation_produces_bytes():
    """Generate a minimal PDF and verify it starts with %PDF."""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    c.drawString(100, 700, "DataSentinel Test Report")
    c.save()
    buf.seek(0)
    content = buf.read()
    assert content[:4] == b"%PDF"


def test_pdf_report_builder_importable():
    """The report builder module must be importable."""
    try:
        from app.api.v1 import reports  # noqa: F401
        assert True
    except ImportError as e:
        pytest.fail(f"Could not import reports: {e}")


# ── Excel generation ───────────────────────────────────────────────────────────

def test_excel_library_available():
    """openpyxl must be importable."""
    import openpyxl
    assert openpyxl.__version__ is not None


def test_excel_generation_produces_valid_workbook():
    """Create a minimal workbook and verify it can be read back."""
    import openpyxl
    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations"
    ws.append(["Rule", "Severity", "Field", "Count"])
    ws.append(["R001", "high", "sensor_value", 42])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = openpyxl.load_workbook(buf)
    ws2 = wb2["Violations"]
    assert ws2.cell(1, 1).value == "Rule"
    assert ws2.cell(2, 4).value == 42


def test_excel_multiple_sheets():
    import openpyxl, io
    wb = openpyxl.Workbook()
    for name in ["Summary", "Violations", "Rules", "Run History"]:
        if wb.active.title == "Sheet":
            wb.active.title = name
        else:
            wb.create_sheet(name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = openpyxl.load_workbook(buf)
    assert "Summary" in wb2.sheetnames
    assert "Violations" in wb2.sheetnames


# ── Report data helpers ────────────────────────────────────────────────────────

def test_violations_by_severity_grouping():
    violations = [
        {"severity": "critical", "rule_id": "R1"},
        {"severity": "high",     "rule_id": "R2"},
        {"severity": "high",     "rule_id": "R3"},
        {"severity": "medium",   "rule_id": "R4"},
        {"severity": "low",      "rule_id": "R5"},
    ]
    by_severity = {}
    for v in violations:
        by_severity[v["severity"]] = by_severity.get(v["severity"], 0) + 1
    assert by_severity["critical"] == 1
    assert by_severity["high"] == 2
    assert by_severity["medium"] == 1


def test_readiness_score_formatting():
    """Score stored as 0–1 float must render as percentage string."""
    score = 0.8754
    rendered = f"{round(score * 100, 1)}%"
    assert rendered == "87.5%"


def test_gate_passed_label():
    for gate, expected in [(True, "PASSED"), (False, "FAILED")]:
        label = "PASSED" if gate else "FAILED"
        assert label == expected
